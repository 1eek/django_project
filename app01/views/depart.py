
from django.shortcuts import render, redirect
from app01 import models


# Create your views here.

def depart_list(req):
    # 去数据库中获取所有的部门列表
    queryset = models.Department.objects.all()
    return render(req, "depart_list.html", {'queryset': queryset})


def depart_add(req):
    """添加部门"""
    if (req.method == "GET"):
        return render(req, "depart_add.html")
    title = req.POST.get("title")
    models.Department.objects.create(title=title)
    return redirect("/depart/list/")


def depart_delete(req):
    nid = req.GET.get('nid')
    models.Department.objects.filter(id=nid).delete()
    return redirect("/depart/list/")


def depart_edit(req, nid):
    """编辑部门"""
    if (req.method == "GET"):
        row_object = models.Department.objects.filter(id=nid).first()
        return render(req, "depart_edit.html", {'row_object': row_object})
    title = req.POST.get('title')
    models.Department.objects.filter(id=nid).update(title=title)
    return redirect("/depart/list")

from openpyxl import load_workbook
def depart_multi(req):
    """批量上传基于Excel"""
    #获取用户上传的文件对象
    file_obj = req.Files.get("exc")
    #读取文件内容
    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]
    #循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0]
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title = text)

        return redirect('/depart/list/')
